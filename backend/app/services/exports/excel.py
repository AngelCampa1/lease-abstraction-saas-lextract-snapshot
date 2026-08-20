"""Excel workbook exporter using openpyxl.

Generates a professional .xlsx lease abstraction report with a summary
sheet and 14 category sheets, each containing field/value/confidence
tables with conditional formatting.
"""

from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.services.brand import BRAND_ASSETS
from app.services.exports.base import EXPORT_DISCLAIMER, ExportBase
from app.services.exports.templates import (
    DEFAULT_TEMPLATE,
    FIELD_LABELS,
    TEMPLATES,
    TemplateSection,
)
from app.services.exports.utils import (
    field_confidence_scores,
    format_value,
    get_confidence_tier,
    unwrap_field,
)

# Confidence tier fills
CONFIDENCE_FILLS: dict[str, PatternFill] = {
    "HIGH": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "MEDIUM": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    "LOW": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    "N/A": PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid"),
}

HEADER_FILL = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
_SPREADSHEET_FORMULA_PREFIXES = ("=", "+", "-", "@", "\t", "\r", "\n")

_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _safe_cell_text(value: Any) -> str:
    """Format a value for Excel while forcing user content to remain text."""
    text = format_value(value)
    if not text:
        return text
    visible = text.lstrip(" \t\r\n")
    if text.startswith(_SPREADSHEET_FORMULA_PREFIXES) or visible.startswith(
        _SPREADSHEET_FORMULA_PREFIXES
    ):
        return f"'{text}"
    return text


class ExcelExporter(ExportBase):
    """Generate Excel (.xlsx) export of extraction results.

    Produces a workbook with a summary sheet and 14 category data sheets,
    each with field/value/confidence tables and conditional fill colors.
    """

    @property
    def content_type(self) -> str:
        """MIME type for Excel documents."""
        return "application/vnd.openxmlformats-officedocument" ".spreadsheetml.sheet"

    @property
    def extension(self) -> str:
        """File extension for Excel documents."""
        return "xlsx"

    def generate(
        self,
        extraction_data: dict[str, Any],
        confidence_scores: dict[str, Any],
        red_flags: list[dict[str, Any]],
        template: str,
        document_filename: str,
    ) -> bytes:
        """Generate an Excel workbook export.

        Args:
            extraction_data: Extracted lease fields keyed by field_name.
            confidence_scores: Confidence tier per field.
            red_flags: List of detected red flag dicts.
            template: Template name (e.g. 'commercial', 'office').
            document_filename: Original uploaded PDF filename.

        Returns:
            Raw bytes of the generated .xlsx file.
        """
        wb = Workbook()
        # Use the default active sheet as Summary
        active_sheet = wb.active
        if not isinstance(active_sheet, Worksheet):
            raise RuntimeError("Excel workbook did not create a worksheet")
        summary_ws = active_sheet
        summary_ws.title = "Summary"
        self._add_summary_sheet(
            summary_ws,
            extraction_data,
            confidence_scores,
            red_flags,
            document_filename,
        )

        # Create category sheets
        template_config = TEMPLATES.get(template, TEMPLATES[DEFAULT_TEMPLATE])
        for section in template_config.sections:
            ws: Worksheet = wb.create_sheet(title=section.display_name[:31])
            self._add_category_sheet(ws, section, extraction_data, confidence_scores)

        if red_flags:
            red_flags_ws: Worksheet = wb.create_sheet(title="Appendix - Red Flags")
            self._add_red_flags_appendix_sheet(red_flags_ws, red_flags)

        buffer = BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    def _add_summary_sheet(
        self,
        ws: Worksheet,
        extraction_data: dict[str, Any],
        confidence_scores: dict[str, Any],
        red_flags: list[dict[str, Any]],
        document_filename: str,
    ) -> None:
        """Populate the summary sheet with overview data.

        Args:
            ws: The worksheet to populate.
            extraction_data: Extracted lease fields.
            confidence_scores: Confidence tier per field.
            red_flags: List of detected red flags.
            document_filename: Original PDF filename.
        """
        logo = Image(str(BRAND_ASSETS.logo_path))
        logo.width = 220
        logo.height = 57
        ws.add_image(logo, "A1")

        # Row 5: Title
        title_cell = ws.cell(row=5, column=1, value="Lease Abstraction Report")
        title_cell.font = Font(name="Calibri", size=16, bold=True)

        # Row 6: Filename
        ws.cell(row=6, column=1, value=_safe_cell_text(document_filename))

        # Row 8-10: Key identifiers
        landlord = _safe_cell_text(
            unwrap_field(extraction_data.get("landlord_legal_name"))
        )
        tenant = _safe_cell_text(unwrap_field(extraction_data.get("tenant_legal_name")))
        premises = _safe_cell_text(
            unwrap_field(extraction_data.get("premises_address"))
        )

        ws.cell(row=8, column=1, value="Landlord")
        ws.cell(row=8, column=2, value=landlord)
        ws.cell(row=9, column=1, value="Tenant")
        ws.cell(row=9, column=2, value=tenant)
        ws.cell(row=10, column=1, value="Premises")
        ws.cell(row=10, column=2, value=premises)

        for row_idx in range(8, 11):
            ws.cell(row=row_idx, column=1).font = Font(
                name="Calibri", size=10, bold=True
            )

        # Row 12: Confidence Distribution header
        conf_header = ws.cell(row=12, column=1, value="Confidence Distribution")
        conf_header.font = Font(name="Calibri", size=12, bold=True)

        # Row 13-16: Confidence counts
        field_scores = field_confidence_scores(confidence_scores)
        high_count = sum(
            1 for v in field_scores.values() if get_confidence_tier(v) == "HIGH"
        )
        medium_count = sum(
            1 for v in field_scores.values() if get_confidence_tier(v) == "MEDIUM"
        )
        low_count = sum(
            1 for v in field_scores.values() if get_confidence_tier(v) == "LOW"
        )
        na_count = sum(
            1 for v in field_scores.values() if get_confidence_tier(v) == "N/A"
        )

        ws.cell(row=13, column=1, value="HIGH")
        ws.cell(row=13, column=2, value=high_count)
        ws.cell(row=13, column=1).fill = CONFIDENCE_FILLS["HIGH"]

        ws.cell(row=14, column=1, value="MEDIUM")
        ws.cell(row=14, column=2, value=medium_count)
        ws.cell(row=14, column=1).fill = CONFIDENCE_FILLS["MEDIUM"]

        ws.cell(row=15, column=1, value="LOW")
        ws.cell(row=15, column=2, value=low_count)
        ws.cell(row=15, column=1).fill = CONFIDENCE_FILLS["LOW"]

        ws.cell(row=16, column=1, value="N/A (not in lease)")
        ws.cell(row=16, column=2, value=na_count)
        ws.cell(row=16, column=1).fill = CONFIDENCE_FILLS["N/A"]

        # Row 18: Red flags
        ws.cell(
            row=18,
            column=1,
            value=f"Red Flags: {len(red_flags)} detected",
        )
        ws.cell(row=18, column=1).font = Font(name="Calibri", size=10, bold=True)

        # Row 20: AI-accuracy disclaimer
        disclaimer_cell = ws.cell(row=20, column=1, value=EXPORT_DISCLAIMER)
        disclaimer_cell.font = Font(name="Calibri", size=8, italic=True, color="595959")
        disclaimer_cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=20, start_column=1, end_row=20, end_column=2)

        # Auto-width columns
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 40

    def _add_category_sheet(
        self,
        ws: Worksheet,
        section: TemplateSection,
        extraction_data: dict[str, Any],
        confidence_scores: dict[str, Any],
    ) -> None:
        """Populate a category sheet with field data.

        Args:
            ws: The worksheet to populate.
            section: Template section definition.
            extraction_data: Extracted lease fields.
            confidence_scores: Confidence tier per field.
        """
        headers = ["Field Label", "Value", "Confidence Score", "Confidence Tier"]

        # Header row
        for col_idx, header_text in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header_text)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center")
            cell.border = _THIN_BORDER

        # Data rows
        for row_offset, field_name in enumerate(section.fields, start=2):
            raw_value = extraction_data.get(field_name)
            value = unwrap_field(raw_value)
            conf_data = confidence_scores.get(field_name)
            tier = get_confidence_tier(conf_data)
            label = FIELD_LABELS.get(field_name, field_name)

            # Extract numeric score if available
            score: float | None = None
            if isinstance(conf_data, dict) and "score" in conf_data:
                score = float(conf_data["score"])

            # Field Label
            label_cell = ws.cell(row=row_offset, column=1, value=label)
            label_cell.font = Font(name="Calibri", size=10, bold=True)
            label_cell.border = _THIN_BORDER

            # Value
            value_cell = ws.cell(row=row_offset, column=2, value=_safe_cell_text(value))
            value_cell.border = _THIN_BORDER

            # Confidence Score
            score_cell = ws.cell(
                row=row_offset,
                column=3,
                value=score if score is not None else "",
            )
            score_cell.border = _THIN_BORDER
            if score is not None:
                score_cell.number_format = "0.00"

            # Confidence Tier with fill
            tier_cell = ws.cell(row=row_offset, column=4, value=tier if tier else "")
            tier_cell.border = _THIN_BORDER
            if tier in CONFIDENCE_FILLS:
                tier_cell.fill = CONFIDENCE_FILLS[tier]

        # Auto-width columns
        for col_idx in range(1, 5):
            col_letter = get_column_letter(col_idx)
            max_length = len(headers[col_idx - 1])
            for row_idx in range(2, len(section.fields) + 2):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value is not None:
                    max_length = max(max_length, len(str(cell_value)))
            ws.column_dimensions[col_letter].width = min(max_length + 4, 50)

    def _add_red_flags_appendix_sheet(
        self,
        ws: Worksheet,
        red_flags: list[dict[str, Any]],
    ) -> None:
        """Populate the red flags appendix sheet."""
        title = ws.cell(row=1, column=1, value="Appendix: Red Flags")
        title.font = Font(name="Calibri", size=16, bold=True)

        ws.cell(
            row=2,
            column=1,
            value=f"{len(red_flags)} potential issue(s) were identified.",
        )

        headers = ["Field", "Severity", "Description"]
        for col_idx, header_text in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col_idx, value=header_text)
            cell.font = HEADER_FONT
            cell.fill = PatternFill(
                start_color="8B0000",
                end_color="8B0000",
                fill_type="solid",
            )
            cell.alignment = Alignment(horizontal="center")
            cell.border = _THIN_BORDER

        for row_idx, flag in enumerate(red_flags, start=4):
            field_name = str(flag.get("name", "Unknown"))
            label = _safe_cell_text(FIELD_LABELS.get(field_name, field_name))
            severity = _safe_cell_text(flag.get("severity", "Unknown"))
            description = _safe_cell_text(flag.get("description", ""))

            ws.cell(row=row_idx, column=1, value=label)
            ws.cell(row=row_idx, column=2, value=severity)
            ws.cell(row=row_idx, column=3, value=description)
            for col_idx in range(1, 4):
                ws.cell(row=row_idx, column=col_idx).border = _THIN_BORDER

        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 14
        ws.column_dimensions["C"].width = 80
